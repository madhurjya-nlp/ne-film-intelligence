#!/usr/bin/env python3
"""Create categorized film/media education programs xlsx for Assam/NE India filmmakers.
Cost bands based on estimated tuition (INR lacs). Living, travel, visa separate.
Data compiled June 2026 from official sources + portals. Verify always.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")
band0_fill = PatternFill("solid", fgColor="C6EFCE")  # green 0-3
band1_fill = PatternFill("solid", fgColor="FFEB9C")  # yellow 4-6
band2_fill = PatternFill("solid", fgColor="FFC7CE")  # red-ish 7-9
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical="top")

programs = [
    # 0-3 Lacs
    {"band": "0-3 Lacs", "name": "Dr. Bhupen Hazarika Regional Govt Film & TV Institute (DBHRGFTI)", "country": "Assam, India (Guwahati)", "program": "3yr Diploma Cinematography / Editing / Audiography + Acting Cert", "duration": "3 years / cert", "tuition_lacs": "~0.5-2 (govt subsidized, very low)", "notes": "ONLY govt film institute in NE India. Practical. Start here for Assam filmmakers. Local access from North Lakhimpur. Portfolio builder.", "link": "dbhrgfti.assam.gov.in"},
    {"band": "0-3 Lacs", "name": "Bir Tikendrajit University", "country": "Manipur, India (NE)", "program": "MA Film & Television (Online)", "duration": "2 years", "tuition_lacs": "~0.46 total", "notes": "Extremely affordable, regional NE option. Online flexible. Good foundation degree while working.", "link": "birtikendrajituniversity.ac.in"},
    {"band": "0-3 Lacs", "name": "Sapientia Hungarian University of Transylvania", "country": "Romania", "program": "MA Film Studies", "duration": "2 years", "tuition_lacs": "~0.5-2 per yr (earlier €550/yr reported)", "notes": "Hidden gem. Theory + studies. Very low. On-campus but cheap living. Contact for thesis flexibility.", "link": "sapientia.ro / film.sapientia.ro"},
    {"band": "0-3 Lacs", "name": "Various German Public Universities (Media/Film Studies)", "country": "Germany", "program": "MA Media Studies / Film & TV Studies (select English)", "duration": "2 years", "tuition_lacs": "0 tuition + €150-350/sem admin", "notes": "Tuition free at public. Strong theory, some production. Living 8-12L/yr separate. Good with NOS or self. English MAs increasing.", "link": "study-in-germany.de or mastersportal"},
    {"band": "0-3 Lacs", "name": "Digital Film School Africa (via atingi)", "country": "Pan-Africa / Online", "program": "Full Filmmaking Curriculum (Cinematography, Doc, Editing etc)", "duration": "Self-paced modules", "tuition_lacs": "FREE (certificates)", "notes": "Professional. Accessible. Great stack with local work. Certificates recognized for portfolio.", "link": "digitalfilmschool.africa / atingi.org"},
    {"band": "0-3 Lacs", "name": "EICTV Cuba Online Workshops", "country": "Cuba (online)", "program": "Documentary / Filmmaking Workshops", "duration": "Short (weeks)", "tuition_lacs": "0.4-1.2 per workshop", "notes": "Prestigious (Gabriel Garcia Marquez founded). Online options. Excellent for doc makers from NE.", "link": "eictv.org"},
    {"band": "0-3 Lacs", "name": "Google Skillshop / Adobe Student / Coursera (aid)", "country": "Global / Online", "program": "YouTube Certs, Creative Cloud access, Film Storytelling Specializations", "duration": "Self-paced", "tuition_lacs": "0-0.3 (aid)", "notes": "Stack these immediately. Free Adobe for Indian students via inst. Real value for content + portfolio.", "link": "skillshop.exceedlms.com / coursera.org"},
    {"band": "0-3 Lacs", "name": "Asia e University (AeU)", "country": "Malaysia (online options)", "program": "MSc Multimedia / Media related", "duration": "1.5-2 years", "tuition_lacs": "~2-4 total est (flexible)", "notes": "ODL friendly for Indian students. Malaysia India ties. Check current media programs.", "link": "aeu.edu.my"},
    # 4-6 Lacs
    {"band": "4-6 Lacs", "name": "Satyajit Ray Film & TV Institute (SRFTI)", "country": "India (Kolkata)", "program": "MFA Cinema (various)", "duration": "2 years", "tuition_lacs": "~3-5+ (tuition ~48k/sem + one-time ~80k)", "notes": "Top Indian institute. Competitive entrance (JET). Excellent for serious filmmakers. Closer than abroad.", "link": "srfti.ac.in"},
    {"band": "4-6 Lacs", "name": "FTII Pune (select programs)", "country": "India (Pune)", "program": "PG Diploma / MFA equivalents", "duration": "1-3 years", "tuition_lacs": "~4.85-7 total reported academic", "notes": "Prestigious. High competition. Good national recognition. Consider for mid budget.", "link": "ftii.ac.in"},
    {"band": "4-6 Lacs", "name": "Polish Public Universities (e.g. select Lodz or others)", "country": "Poland", "program": "MA Film / Media / Audiovisual", "duration": "1.5-2 years", "tuition_lacs": "~3-6 total est (€2-6k/yr)", "notes": "English options growing. Strong Eastern European cinema tradition. Affordable living.", "link": "studyinpoland.pl / lodz film school"},
    {"band": "4-6 Lacs", "name": "Tallinn University BFM (select)", "country": "Estonia", "program": "MA Documentary / Cross-Media", "duration": "2 years", "tuition_lacs": "~4-8 total est", "notes": "Part of Erasmus consortia. Baltic network strong. English. Check scholarships.", "link": "tlu.ee / bfm"},
    {"band": "4-6 Lacs", "name": "Lithuanian Academy (LMTA / part consortia)", "country": "Lithuania", "program": "MA Film", "duration": "2 years", "tuition_lacs": "~3-6", "notes": "DocNomads partner. Practical. Affordable Baltic option.", "link": "lmta.lt"},
    # 7-9 Lacs
    {"band": "7-9 Lacs", "name": "FAMU (English programs)", "country": "Czech Republic", "program": "MA Filmmaking (various specializations)", "duration": "2-3 years", "tuition_lacs": "~8-18 total (English higher; Czech free for some)", "notes": "World class. English track expensive. Czech language track cheaper. Legendary alumni.", "link": "famu.cz / international.famu.cz"},
    {"band": "7-9 Lacs", "name": "Arts University Bournemouth or Falmouth (Online)", "country": "UK (online)", "program": "MA Film Practice / Film & TV", "duration": "2 years part-time", "tuition_lacs": "~9-12 but with aid/scholar or NOS reachable lower", "notes": "High quality online. Use NOS (QS ranked) for full funding. Otherwise stretch.", "link": "aub.ac.uk / falmouth.ac.uk"},
    {"band": "7-9 Lacs", "name": "KinoEyes / DocNomads (self-funded)", "country": "Multi Europe (Portugal/Estonia/etc)", "program": "Joint MA Fiction / Documentary", "duration": "2 years", "tuition_lacs": "~10-15 self (0 with Erasmus Mundus)", "notes": "Gold standard. Apply for scholarship primarily. Self-funded in 7-9+ range.", "link": "kinoeyes.eu / docnomads.eu"},
    {"band": "7-9 Lacs", "name": "University of Georgia (Georgian-American Film Academy) or GIPA", "country": "Georgia", "program": "MA Filmmaking / Multimedia Journalism", "duration": "1.5-2 years", "tuition_lacs": "~4-7 est but with living total mid-high", "notes": "Very affordable country. Vibrant indie scene. English programs. Good stepping for Caucasus/Asia networks.", "link": "ug.edu.ge / gipa.ge"},
    {"band": "7-9 Lacs", "name": "Additional: select programs in Hungary / Italy / Spain (English)", "country": "Hungary/Italy/Spain", "program": "MA Film/Media (e.g. ELTE, Roma Tre IMACS, various)", "duration": "1-2 years", "tuition_lacs": "~5-9", "notes": "Varies widely. Public lower end. Check current English film production tracks. Festival access bonus.", "link": "mastersportal.com filter Film + country + tuition"},
]

# Sheet 1: All Categorized
ws = wb.active
ws.title = "Categorized Programs"

headers = ["Cost Band (Tuition)", "Institution", "Country", "Program", "Duration", "Est Tuition (INR Lacs)", "Notes for NE/Assam Filmmaker (Ground Reality)", "Link / Source"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center")
    cell.border = thin_border

for row_idx, p in enumerate(programs, 2):
    ws.cell(row=row_idx, column=1, value=p["band"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=p["name"]).border = thin_border
    ws.cell(row=row_idx, column=3, value=p["country"]).border = thin_border
    ws.cell(row=row_idx, column=4, value=p["program"]).border = thin_border
    ws.cell(row=row_idx, column=5, value=p["duration"]).border = thin_border
    ws.cell(row=row_idx, column=6, value=p["tuition_lacs"]).border = thin_border
    ws.cell(row=row_idx, column=7, value=p["notes"]).border = thin_border
    ws.cell(row=row_idx, column=7).alignment = wrap_align
    ws.cell(row=row_idx, column=8, value=p["link"]).border = thin_border
    if "0-3" in p["band"]:
        ws.cell(row=row_idx, column=1).fill = band0_fill
    elif "4-6" in p["band"]:
        ws.cell(row=row_idx, column=1).fill = band1_fill
    else:
        ws.cell(row=row_idx, column=1).fill = band2_fill

# Column widths
widths = [14, 45, 22, 45, 14, 22, 70, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 25
for r in range(2, len(programs)+2):
    ws.row_dimensions[r].height = 45

# Sheet 2: Summary by Band + Strategy
ws2 = wb.create_sheet("Band Summary + Strategy")
summary = [
    ["Band", "What it covers (tuition est.)", "Recommended for Assam/NE profile", "Priority Actions"],
    ["0-3 Lacs", "Local NE diplomas (DBHRGFTI), cheap online India/Asia, free certs, ultra-low Europe public (Germany/Sapientia), workshops", "Start immediately. Build portfolio locally while stacking certs. DBHRGFTI is ground-zero advantage.", "1. Apply/enroll DBHRGFTI or local. 2. Free MOOCs + Google/Adobe. 3. Make 3-5 shorts. 4. NOS + Erasmus apps."],
    ["4-6 Lacs", "Top Indian institutes (SRFTI/FTII), mid Europe (Poland, Baltic, Georgia)", "After foundation/portfolio. Or direct with partial scholarship/NOS backup.", "Target SRFTI JET + Polish/Georgian as realistic funded or self options."],
    ["7-9 Lacs", "Prestigious English programs (FAMU English, UK online, full consortia self-funded)", "With NOS (full cover) or education loan + strong portfolio. Or strong scholarship.", "Primary: Erasmus Mundus full. Backup: FAMU English or UK online if funded."],
]
for r_idx, row in enumerate(summary, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 55
ws2.column_dimensions['D'].width = 55

# Sheet 3: Key Notes Ground Reality
ws3 = wb.create_sheet("Ground Reality Notes Assam-NE")
notes = [
    ["Topic", "Key Facts for Student/Filmmaker from Assam (North Lakhimpur etc.)"],
    ["NOS Scholarship (ST)", "Ministry Tribal Affairs. Portal overseas.tribal.gov.in . Deadline ~30 June 2026 for 2026-27 (check current). Covers tuition, living (~$15k/yr), travel, insurance. Age <32/35 MA. Family income <=6L. Top QS unis preferred. Film/Media eligible under Humanities/Fine Arts. NO bond. Apply with unconditional offer or in process."],
    ["Local Start", "DBHRGFTI in/near Guwahati: accessible by train from Lakhimpur (6-10 hrs). Build real skills + network at Jyoti Chitraban studio (festivals, industry). Low cost, govt recognized."],
    ["Travel & Visa from Assam", "Guwahati (GAU) airport for most intl. VFS/embassy often Delhi or Kolkata - add travel cost/time. Student visa needs: offer letter, funds proof (scholar helps), medical, ties to India. Start paperwork early."],
    ["Docs & Attestation", "Degree/marksheets attestation: Guwahati HRD/ MEA process. ST certificate critical (keep updated). Passport from Guwahati PSK possible."],
    ["Language & Culture Fit", "Strong English needed (IELTS 6.5+). Assamese stories unique selling point abroad (indigenous, biodiversity, rivers, Bihu, oral trads underrepresented). Subtitles key for Assamese work."],
    ["Costs Beyond Tuition", "Living Europe 8-15L/yr. Flights GAU-Europe 50-90k one way. Health ins ~20-40k/yr. Add 20-30% buffer. Part-time allowed in many EU (10-20h/wk)."],
    ["Recognition & Return", "Good European/Asian degrees generally recognized (verify WES/Indian univ eq). NOS scholars expected to return/contribute (no strict bond but reputation). Local networks + global degree = strong position in NE/Indian media + intl co-prods."],
    ["Festivals as Pathway", "Submit to NEIFF (Manipur), NIIFF, Shillong IFF, MIFF NE sections, then intl (Clermont, Rotterdam HBF, Busan ACF, DocNomads networks). Festival cred > degree sometimes."],
    ["ST + NE Advantage", "Quotas/scholar priority. Unique voice for funding (Doha, HBF prioritize marginal voices). Assam Film Finance Scheme exists for local projects."],
]
for r_idx, row in enumerate(notes, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 120

# Freeze header
for sheet in [ws, ws2, ws3]:
    sheet.freeze_panes = 'A2'

wb.save("Film_Education_Programs_Categorized_0-9Lacs_Assam_NE.xlsx")
print("XLSX created: Film_Education_Programs_Categorized_0-9Lacs_Assam_NE.xlsx")